import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from datetime import datetime, timedelta
import os
import shutil

# Цвета
FILL_PURPLE = PatternFill(start_color='800080', end_color='800080', fill_type='solid')  # Фиолетовый
FILL_GREEN = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')  # Зеленый
FILL_RED = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')    # Красный
FILL_ORANGE = PatternFill(start_color='FFA500', end_color='FFA500', fill_type='solid')  # Оранжевый
FILL_BLUE = PatternFill(start_color='00BFFF', end_color='00BFFF', fill_type='solid')   # Голубой

# Ввод параметров
file1 = input('Введите путь к первому Excel-файлу (по умолчанию "ОФД.xlsx"): ') or 'ОФД.xlsx'
file2 = input('Введите путь ко второму Excel-файлу (по умолчанию "OLAP.xlsx"): ') or 'OLAP.xlsx'
col1 = input('Введите название столбца с датой/временем в первом файле (по умолчанию "Дата/время"): ') or 'Дата/время'
col2 = input('Введите название столбца с датой во втором файле (по умолчанию "Учетный день"): ') or 'Учетный день'
sum1 = 'Итого'
sum2 = 'Сумма со скидкой, р. Всего'

# Для OLAP.xlsx сумма всегда в столбце E (индекс 4)
sum2_col_idx = 4

date_from = input('Введите начальную дату (ДД.ММ.ГГГГ ЧЧ:ММ): ')
date_to = input('Введите конечную дату (ДД.ММ.ГГГГ ЧЧ:ММ): ')

# Преобразование дат
start_dt = datetime.strptime(date_from, '%d.%m.%Y %H:%M')
end_dt = datetime.strptime(date_to, '%d.%m.%Y %H:%M')

def clean_sum(val):
    if pd.isna(val):
        return None
    try:
        val = str(val).replace(' ', '').replace(',', '.').replace('\xa0', '')
        return round(float(val), 2)
    except Exception:
        return None

def collect_pairs(df, date_col, sum_col, use_time, start_dt, end_dt, propagate_dates=False, sum_col_idx=None, sum_clean_mode='auto'):
    if use_time:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce')
    else:
        df[date_col] = pd.to_datetime(df[date_col], errors='coerce').dt.date
        if propagate_dates:
            df[date_col] = df[date_col].ffill()
    in_range_idxs = []
    out_range_idxs = []
    pairs = []
    for idx, row in df.iterrows():
        dt_val = row[date_col]
        if use_time:
            in_range = (isinstance(dt_val, pd.Timestamp) and not pd.isna(dt_val) and start_dt <= dt_val <= end_dt)
        else:
            in_range = (dt_val is not None and not pd.isna(dt_val) and start_dt.date() <= dt_val <= end_dt.date())
        if in_range:
            if sum_col_idx is not None:
                sum_val = row.iloc[sum_col_idx]
            else:
                sum_val = row[sum_col]
            # Разная обработка для файлов
            if sum_clean_mode == 'clean':
                sum_val = clean_sum(sum_val)
            elif sum_clean_mode == 'float':
                try:
                    sum_val = round(float(sum_val), 2)
                except Exception:
                    sum_val = None
            else:
                # авто: если строка, чистим, если число — просто округляем
                if isinstance(sum_val, str):
                    sum_val = clean_sum(sum_val)
                else:
                    try:
                        sum_val = round(float(sum_val), 2)
                    except Exception:
                        sum_val = None
            if sum_val is None:
                continue
            pairs.append((dt_val, sum_val, idx))
            in_range_idxs.append(idx)
        else:
            out_range_idxs.append(idx)
    return pairs, out_range_idxs

# --- Функция для окрашивания строк ---
def color_rows(filepath, header_row, pairs_to_color, color, row_offset=2):
    wb = load_workbook(filepath)
    ws = wb.active
    for idx in pairs_to_color:
        excel_row = header_row + 2 + idx  # +2: 1 для смещения к данным, 1 для индексации с 1
        for cell in ws[excel_row]:
            cell.fill = color
    wb.save(filepath)

# --- Основная логика ---
def main():
    df1 = pd.read_excel(file1, header=0)
    df2 = pd.read_excel(file2, header=3)
    pairs1, out1 = collect_pairs(df1, col1, sum1, use_time=True, start_dt=start_dt, end_dt=end_dt, sum_clean_mode='clean')
    pairs2, out2 = collect_pairs(df2, col2, sum2, use_time=False, start_dt=start_dt, end_dt=end_dt, propagate_dates=True, sum_col_idx=sum2_col_idx, sum_clean_mode='float')
    used1 = set()
    used2 = set()
    green1 = []
    green2 = []
    red1 = []
    red2 = []
    debug_mismatches = []
    for i, (dt1, s1, idx1) in enumerate(pairs1):
        found = False
        for j, (dt2, s2, idx2) in enumerate(pairs2):
            if j in used2:
                continue
            # Сравниваем только по дате
            d1 = dt1.date() if isinstance(dt1, pd.Timestamp) else dt1
            d2 = dt2.date() if isinstance(dt2, pd.Timestamp) else dt2
            if d1 == d2 and abs(s1 - s2) < 0.01:
                green1.append(idx1)
                green2.append(idx2)
                used1.add(i)
                used2.add(j)
                found = True
                break
        if not found:
            red1.append(idx1)
            if len(debug_mismatches) < 5:
                debug_mismatches.append((dt1, s1, '->', [(dt2, s2) for (dt2, s2, _) in pairs2 if (dt2.date() if isinstance(dt2, pd.Timestamp) else dt2) == (dt1.date() if isinstance(dt1, pd.Timestamp) else dt1)]))
    for j, (dt2, s2, idx2) in enumerate(pairs2):
        if j in used2:
            continue
        found = False
        for i, (dt1, s1, idx1) in enumerate(pairs1):
            if i in used1:
                continue
            d1 = dt1.date() if isinstance(dt1, pd.Timestamp) else dt1
            d2 = dt2.date() if isinstance(dt2, pd.Timestamp) else dt2
            if d1 == d2 and abs(s1 - s2) < 0.01:
                found = True
                break
        if not found:
            red2.append(idx2)
    
    # Проверяем столбец "Признак расчета" в ОФД
    orange1 = []
    orange2 = []
    if 'Признак расчета' in df1.columns:
        # Проверяем строки, которые попали в зелёный или красный
        for idx in green1 + red1:
            if idx < len(df1):
                признак = str(df1.iloc[idx]['Признак расчета']).strip()
                if 'Возврат прихода' in признак:
                    orange1.append(idx)
                    # Ищем соответствующую строку в OLAP
                    dt1 = df1.iloc[idx][col1]
                    s1 = clean_sum(df1.iloc[idx][sum1])
                    d1 = dt1.date() if isinstance(dt1, pd.Timestamp) else dt1
                    for j, (dt2, s2, idx2) in enumerate(pairs2):
                        d2 = dt2.date() if isinstance(dt2, pd.Timestamp) else dt2
                        if d1 == d2 and abs(s1 - s2) < 0.01:
                            orange2.append(idx2)
                            break
    
    # Проверяем красные строки на совпадение суммы с разницей в дате до 1 дня
    blue1 = []
    blue2 = []
    used_for_blue1 = set()
    used_for_blue2 = set()
    
    # Проверяем красные строки из первого файла
    for idx1 in red1:
        if idx1 in used_for_blue1:
            continue
        dt1 = df1.iloc[idx1][col1]
        s1 = clean_sum(df1.iloc[idx1][sum1])
        d1 = dt1.date() if isinstance(dt1, pd.Timestamp) else dt1
        
        for j, (dt2, s2, idx2) in enumerate(pairs2):
            if idx2 in used_for_blue2:
                continue
            d2 = dt2.date() if isinstance(dt2, pd.Timestamp) else dt2
            # Проверяем сумму и разницу в дате до 1 дня
            if abs(s1 - s2) < 0.01 and abs((d1 - d2).days) <= 1:
                blue1.append(idx1)
                blue2.append(idx2)
                used_for_blue1.add(idx1)
                used_for_blue2.add(idx2)
                break
    
    # Проверяем красные строки из второго файла
    for idx2 in red2:
        if idx2 in used_for_blue2:
            continue
        dt2 = df2.iloc[idx2][col2]
        s2 = clean_sum(df2.iloc[idx2].iloc[sum2_col_idx])
        d2 = dt2.date() if isinstance(dt2, pd.Timestamp) else dt2
        
        for i, (dt1, s1, idx1) in enumerate(pairs1):
            if idx1 in used_for_blue1:
                continue
            d1 = dt1.date() if isinstance(dt1, pd.Timestamp) else dt1
            # Проверяем сумму и разницу в дате до 1 дня
            if abs(s1 - s2) < 0.01 and abs((d1 - d2).days) <= 1:
                blue1.append(idx1)
                blue2.append(idx2)
                used_for_blue1.add(idx1)
                used_for_blue2.add(idx2)
                break
    
    # Копируем исходные файлы для окрашивания
    out1_file = os.path.splitext(file1)[0] + '_marked.xlsx'
    out2_file = os.path.splitext(file2)[0] + '_marked.xlsx'
    shutil.copyfile(file1, out1_file)
    shutil.copyfile(file2, out2_file)
    # Окрашиваем строки только в копиях
    color_rows(out1_file, header_row=0, pairs_to_color=green1, color=FILL_GREEN)
    color_rows(out1_file, header_row=0, pairs_to_color=red1, color=FILL_RED)
    color_rows(out1_file, header_row=0, pairs_to_color=out1, color=FILL_PURPLE)
    color_rows(out2_file, header_row=3, pairs_to_color=green2, color=FILL_GREEN)
    color_rows(out2_file, header_row=3, pairs_to_color=red2, color=FILL_RED)
    color_rows(out2_file, header_row=3, pairs_to_color=out2, color=FILL_PURPLE)
    # Оранжевый цвет (приоритет над зелёным/красным)
    color_rows(out1_file, header_row=0, pairs_to_color=orange1, color=FILL_ORANGE)
    color_rows(out2_file, header_row=3, pairs_to_color=orange2, color=FILL_ORANGE)
    # Голубой цвет (приоритет над красным)
    color_rows(out1_file, header_row=0, pairs_to_color=blue1, color=FILL_BLUE)
    color_rows(out2_file, header_row=3, pairs_to_color=blue2, color=FILL_BLUE)
    print(f'Обработка завершена. Окрашенные копии файлы: {out1_file}, {out2_file}')
    
    # Статистика по цветам
    print(f'\nСтатистика по первому файлу (ОФД):')
    print(f'  Оранжевых (возврат): {len(orange1)}')
    print(f'  Голубых (±1 день): {len(blue1)}')
    print(f'  Красных (несовпадение): {len(red1)}')
    
    print(f'\nСтатистика по второму файлу (OLAP):')
    print(f'  Оранжевых (возврат): {len(orange2)}')
    print(f'  Голубых (±1 день): {len(blue2)}')
    print(f'  Красных (несовпадение): {len(red2)}')
    
    # Общее количество несоответствий
    total_mismatches = len(orange1) + len(orange2) + len(blue1) + len(blue2) + len(red1) + len(red2)
    print(f'\nОбщее количество несоответствий: {total_mismatches}')
    
    # Создаем файл статистики
    stats_file = "Статистика несоответствий.xlsx"
    
    # Подготавливаем данные для статистики
    stats_data = []
    
    # Общая статистика
    stats_data.append(['ОБЩАЯ СТАТИСТИКА', ''])
    stats_data.append(['Файл ОФД', ''])
    stats_data.append(['Оранжевых (возврат)', len(orange1)])
    stats_data.append(['Голубых (±1 день)', len(blue1)])
    stats_data.append(['Красных (несовпадение)', len(red1)])
    stats_data.append(['', ''])
    stats_data.append(['Файл OLAP', ''])
    stats_data.append(['Оранжевых (возврат)', len(orange2)])
    stats_data.append(['Голубых (±1 день)', len(blue2)])
    stats_data.append(['Красных (несовпадение)', len(red2)])
    stats_data.append(['', ''])
    stats_data.append(['ОБЩЕЕ КОЛИЧЕСТВО НЕСООТВЕТСТВИЙ', total_mismatches])
    stats_data.append(['', ''])
    
    # Статистика по датам для ОФД
    if orange1 or blue1 or red1:
        stats_data.append(['СТАТИСТИКА ПО ДАТАМ - ОФД', ''])
        stats_data.append(['Дата', 'Оранжевые', 'Голубые', 'Красные'])
        
        # Собираем все даты
        all_dates = set()
        if orange1:
            df1_orange = df1.iloc[orange1]
            df1_orange_dates = df1_orange[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df1_orange_dates.value_counts().index)
        if blue1:
            df1_blue = df1.iloc[blue1]
            df1_blue_dates = df1_blue[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df1_blue_dates.value_counts().index)
        if red1:
            df1_red = df1.iloc[red1]
            df1_red_dates = df1_red[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df1_red_dates.value_counts().index)
        
        # Создаем словари для подсчета
        orange_by_date = {}
        blue_by_date = {}
        red_by_date = {}
        
        if orange1:
            df1_orange = df1.iloc[orange1]
            df1_orange_dates = df1_orange[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            orange_by_date = df1_orange_dates.value_counts().to_dict()
        
        if blue1:
            df1_blue = df1.iloc[blue1]
            df1_blue_dates = df1_blue[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            blue_by_date = df1_blue_dates.value_counts().to_dict()
        
        if red1:
            df1_red = df1.iloc[red1]
            df1_red_dates = df1_red[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            red_by_date = df1_red_dates.value_counts().to_dict()
        
        for date in sorted(all_dates):
            orange_count = orange_by_date.get(date, 0)
            blue_count = blue_by_date.get(date, 0)
            red_count = red_by_date.get(date, 0)
            stats_data.append([str(date), orange_count, blue_count, red_count])
        
        stats_data.append(['', '', '', ''])
    
    # Статистика по датам для OLAP
    if orange2 or blue2 or red2:
        stats_data.append(['СТАТИСТИКА ПО ДАТАМ - OLAP', ''])
        stats_data.append(['Дата', 'Оранжевые', 'Голубые', 'Красные'])
        
        # Собираем все даты
        all_dates = set()
        if orange2:
            df2_orange = df2.iloc[orange2]
            df2_orange_dates = df2_orange[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df2_orange_dates.value_counts().index)
        if blue2:
            df2_blue = df2.iloc[blue2]
            df2_blue_dates = df2_blue[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df2_blue_dates.value_counts().index)
        if red2:
            df2_red = df2.iloc[red2]
            df2_red_dates = df2_red[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            all_dates.update(df2_red_dates.value_counts().index)
        
        # Создаем словари для подсчета
        orange_by_date = {}
        blue_by_date = {}
        red_by_date = {}
        
        if orange2:
            df2_orange = df2.iloc[orange2]
            df2_orange_dates = df2_orange[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            orange_by_date = df2_orange_dates.value_counts().to_dict()
        
        if blue2:
            df2_blue = df2.iloc[blue2]
            df2_blue_dates = df2_blue[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            blue_by_date = df2_blue_dates.value_counts().to_dict()
        
        if red2:
            df2_red = df2.iloc[red2]
            df2_red_dates = df2_red[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
            red_by_date = df2_red_dates.value_counts().to_dict()
        
        for date in sorted(all_dates):
            orange_count = orange_by_date.get(date, 0)
            blue_count = blue_by_date.get(date, 0)
            red_count = red_by_date.get(date, 0)
            stats_data.append([str(date), orange_count, blue_count, red_count])
    
    # Создаем DataFrame и сохраняем в Excel
    stats_df = pd.DataFrame(stats_data)
    stats_df.to_excel(stats_file, index=False, header=False)
    print(f'\nСтатистика сохранена в файл: {stats_file}')
    
    # Статистика по датам
    if orange1:
        print('\nОранжевые строки в первом файле (ОФД):')
        df1_orange = df1.iloc[orange1]
        df1_orange_dates = df1_orange[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df1_orange_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')
    
    if orange2:
        print('\nОранжевые строки во втором файле (OLAP):')
        df2_orange = df2.iloc[orange2]
        df2_orange_dates = df2_orange[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df2_orange_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')
    
    if blue1:
        print('\nГолубые строки в первом файле (ОФД):')
        df1_blue = df1.iloc[blue1]
        df1_blue_dates = df1_blue[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df1_blue_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')
    
    if blue2:
        print('\nГолубые строки во втором файле (OLAP):')
        df2_blue = df2.iloc[blue2]
        df2_blue_dates = df2_blue[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df2_blue_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')
    
    if red1:
        print('\nКрасные строки в первом файле (ОФД):')
        df1_red = df1.iloc[red1]
        df1_red_dates = df1_red[col1].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df1_red_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')
    
    if red2:
        print('\nКрасные строки во втором файле (OLAP):')
        df2_red = df2.iloc[red2]
        df2_red_dates = df2_red[col2].apply(lambda x: x.date() if hasattr(x, 'date') else x)
        for date, count in df2_red_dates.value_counts().sort_index().items():
            print(f'  {date}: {count}')

if __name__ == '__main__':
    main() 